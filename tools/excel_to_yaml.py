#!/usr/bin/env python3
import argparse
from pathlib import Path
from typing import Any, Dict, List

import yaml
from openpyxl import load_workbook

from validate_workbook import validate_workbook

TEMPLATE_MAP = {
    "CORE": "core_vsx.j2",
    "8100": "core_vsx.j2",
    "6200F": "access_stack.j2",
    "6300M": "access_stack.j2",
}


def read_sheet(workbook, name: str) -> List[Dict[str, Any]]:
    ws = workbook[name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(cell).strip() for cell in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(row):
            continue
        item = {headers[idx]: value for idx, value in enumerate(row)}
        data.append(item)
    return data


def split_list(value: Any) -> List[str]:
    if not value:
        return []
    if isinstance(value, str):
        return [item.strip() for item in value.split(",") if item.strip()]
    return [str(value)]


def normalize_bool(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return value.strip().upper() in {"Y", "YES", "TRUE"}
    return bool(value)


def normalize_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def write_yaml(path: Path, data: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        yaml.safe_dump(data, handle, sort_keys=False)


def build_inventory(site: Dict[str, Any], devices: List[Dict[str, Any]]) -> Dict[str, Any]:
    inventory = {
        "all": {
            "children": {
                "core": {"hosts": {}},
                "access": {"hosts": {}},
            }
        }
    }
    for device in devices:
        host_entry = {
            "ansible_host": device["mgmt_ip"],
        }
        role = device["role"].lower()
        inventory["all"]["children"][role]["hosts"][device["device_name"]] = host_entry
    return inventory


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert Excel workbook to Ansible inventory.")
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    errors = validate_workbook(args.workbook)
    if errors:
        print("Workbook validation failed:")
        for error in errors:
            print(f"- {error}")
        return 1

    workbook = load_workbook(args.workbook, data_only=True)
    site_rows = read_sheet(workbook, "SITE")
    site = site_rows[0]
    devices = read_sheet(workbook, "DEVICES")
    vlans = read_sheet(workbook, "VLANS")
    core_vsx_rows = read_sheet(workbook, "CORE_VSX")
    access_uplinks = read_sheet(workbook, "ACCESS_UPLINKS")
    interface_desc = (
        read_sheet(workbook, "INTERFACE_DESCRIPTIONS")
        if "INTERFACE_DESCRIPTIONS" in workbook.sheetnames
        else []
    )

    site_vars = {
        "site_name": normalize_str(site["site_name"]),
        "timezone": normalize_str(site["timezone"]),
        "design_mode": normalize_str(site["design_mode"]),
        "default_gateway_mode": normalize_str(site["default_gateway_mode"]),
        "dns_servers": split_list(site["dns_servers"]),
        "ntp_servers": split_list(site["ntp_servers"]),
        "syslog_servers": split_list(site.get("syslog_servers")),
        "mst_region_name": normalize_str(site["mst_region_name"]),
        "mst_revision": int(site["mst_revision"]),
        "mgmt_vlan_id": int(site["mgmt_vlan_id"]),
        "mgmt_vlan_name": normalize_str(site["mgmt_vlan_name"]),
        "mgmt_svi_ip": normalize_str(site["mgmt_svi_ip"]),
        "mgmt_svi_mask": normalize_str(site["mgmt_svi_mask"]),
        "clearpass_enabled": normalize_bool(site["clearpass_enabled"]),
        "clearpass_servers": split_list(site.get("clearpass_servers")),
    }

    vlan_entries = []
    for vlan in vlans:
        vlan_entries.append(
            {
                "vlan_id": int(vlan["vlan_id"]),
                "vlan_name": vlan["vlan_name"],
                "purpose": vlan.get("purpose"),
                "enabled": normalize_bool(vlan["enabled"]),
                "svi_gateway_ip": vlan.get("svi_gateway_ip"),
                "svi_mask": vlan.get("svi_mask"),
                "dhcp_relay_ips": split_list(vlan.get("dhcp_relay_ips")),
            }
        )

    core_vsx = {row["vsx_pair_id"]: row for row in core_vsx_rows}
    uplink_map = {row["stack_name"]: row for row in access_uplinks}

    output_root = args.output
    output_root.mkdir(parents=True, exist_ok=True)

    device_names = [normalize_str(device["device_name"]) for device in devices]
    duplicates = {name for name in device_names if device_names.count(name) > 1}
    if duplicates:
        print(f"Duplicate device_name entries found: {', '.join(sorted(duplicates))}")
        return 1

    inventory = build_inventory(site_vars, devices)
    write_yaml(output_root / "hosts.yml", inventory)

    group_vars = site_vars.copy()
    group_vars["vlans"] = vlan_entries
    group_vars["access_uplinks"] = access_uplinks
    write_yaml(output_root / "group_vars" / "all.yml", group_vars)

    core_devices = [
        normalize_str(device["device_name"])
        for device in devices
        if normalize_str(device["role"]).lower() == "core"
    ]
    access_devices = [
        normalize_str(device["device_name"])
        for device in devices
        if normalize_str(device["role"]).lower() == "access"
    ]
    write_yaml(
        output_root / "group_vars" / "core.yml",
        {"role": "core", "device_names": core_devices},
    )
    write_yaml(
        output_root / "group_vars" / "access.yml",
        {"role": "access", "device_names": access_devices},
    )

    interface_desc_map: Dict[str, List[Dict[str, Any]]] = {}
    for item in interface_desc:
        interface_desc_map.setdefault(normalize_str(item["device_name"]), []).append(
            {
                "interface": normalize_str(item["interface"]),
                "description": normalize_str(item["description"]),
            }
        )

    for device in devices:
        device_name = normalize_str(device["device_name"])
        role = normalize_str(device["role"]).lower()
        model = normalize_str(device["model"])
        host_vars: Dict[str, Any] = {
            "device_name": device_name,
            "role": role,
            "model": model,
            "mgmt_ip": normalize_str(device["mgmt_ip"]),
            "mgmt_mask": normalize_str(device["mgmt_mask"]),
            "mgmt_gateway": normalize_str(device["mgmt_gateway"]),
            "template": TEMPLATE_MAP.get(model, ""),
            "interface_descriptions": interface_desc_map.get(device_name, []),
        }

        if role == "core":
            host_vars.update(
                {
                    "vsx_pair_id": normalize_str(device.get("vsx_pair_id")),
                    "vsx_role": normalize_str(device.get("vsx_role")),
                    "core_vsx": core_vsx.get(normalize_str(device.get("vsx_pair_id")), {}),
                }
            )
        else:
            host_vars.update(
                {
                    "stack_name": normalize_str(device.get("stack_name")),
                    "stack_member_id": device.get("stack_member_id"),
                    "uplink": uplink_map.get(normalize_str(device.get("stack_name")), {}),
                }
            )

        write_yaml(output_root / "host_vars" / f"{device_name}.yml", host_vars)

    print(f"Inventory written to {output_root}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
