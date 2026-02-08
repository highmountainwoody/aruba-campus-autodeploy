#!/usr/bin/env python3
import argparse
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

REQUIRED_TABS = {
    "SITE": [
        "site_name",
        "timezone",
        "design_mode",
        "default_gateway_mode",
        "dns_servers",
        "ntp_servers",
        "syslog_servers",
        "mst_region_name",
        "mst_revision",
        "mgmt_vlan_id",
        "mgmt_vlan_name",
        "mgmt_svi_ip",
        "mgmt_svi_mask",
        "clearpass_enabled",
        "clearpass_servers",
    ],
    "DEVICES": [
        "device_name",
        "role",
        "model",
        "mgmt_ip",
        "mgmt_mask",
        "mgmt_gateway",
        "vsx_pair_id",
        "vsx_role",
        "stack_name",
        "stack_member_id",
    ],
    "VLANS": [
        "vlan_id",
        "vlan_name",
        "purpose",
        "enabled",
        "svi_ip",
        "svi_mask",
        "dhcp_relay_ips",
    ],
    "CORE_VSX": ["vsx_pair_id", "isl_port_1", "isl_port_2"],
    "ACCESS_UPLINKS": [
        "stack_name",
        "uplink_port_1",
        "uplink_port_2",
        "lacp_group_id",
        "core_peer_1",
        "core_peer_1_port",
        "core_peer_1_ip",
        "access_peer_1_ip",
        "core_peer_2",
        "core_peer_2_port",
        "core_peer_2_ip",
        "access_peer_2_ip",
    ],
}

OPTIONAL_TABS = {
    "INTERFACE_DESCRIPTIONS": ["device_name", "interface", "description"],
}


def normalize_headers(headers: List[str]) -> List[str]:
    return [str(header).strip() if header is not None else "" for header in headers]


def tab_headers(ws) -> List[str]:
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    return normalize_headers(headers)


def validate_tab(ws, required_columns: List[str]) -> List[str]:
    errors = []
    headers = tab_headers(ws)
    missing = [col for col in required_columns if col not in headers]
    if missing:
        errors.append(
            f"Missing columns in tab '{ws.title}': {', '.join(missing)}"
        )
    return errors


def validate_site_tab(ws) -> List[str]:
    errors = []
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows or all(value is None for value in rows[0]):
        errors.append("SITE tab must have exactly one populated row.")
        return errors
    if len([row for row in rows if any(row)]) > 1:
        errors.append("SITE tab must contain exactly one row of data.")
    return errors


def validate_workbook(path: Path) -> List[str]:
    errors: List[str] = []
    workbook = load_workbook(path, data_only=True)

    for tab, columns in REQUIRED_TABS.items():
        if tab not in workbook.sheetnames:
            errors.append(f"Missing required tab: {tab}")
            continue
        errors.extend(validate_tab(workbook[tab], columns))
        if tab == "SITE":
            errors.extend(validate_site_tab(workbook[tab]))

    for tab, columns in OPTIONAL_TABS.items():
        if tab in workbook.sheetnames:
            errors.extend(validate_tab(workbook[tab], columns))

    return errors


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Validate Aruba CX campus Excel workbook."
    )
    parser.add_argument("--workbook", required=True, type=Path)
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"Workbook not found: {args.workbook}")
        return 2

    errors = validate_workbook(args.workbook)
    if errors:
        print("Workbook validation failed:")
        for error in errors:
            print(f"- {error}")
        return 1

    print("Workbook validation passed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
